#!/usr/bin/env python3
"""
build_sheet.py — English Open 2026 fixture workbook

Reads fixtures.csv and writes english-open-2026.xlsx: an Overview tab, one tab
per day, and one tab per tracked player. Upload to Drive with conversion on and
it becomes a native Google Sheet with the tabs and formatting intact.

Run:  python build_sheet.py
"""

import csv
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CSV_FILE = "fixtures.csv"
OUTPUT_FILE = "english-open-2026-full.xlsx"
COMPACT_FILE = "english-open-2026.xlsx"

# Pickleball England red, with the softer tints used for banding and accents.
BRAND = "A6192E"
INK = "1F2933"
MUTED = "7B8794"
BAND = "F7F8FA"
RULE = "E4E7EB"
FLAG = "FFF4E5"

# One accent per player, used on their own tab header and the Overview grid.
PLAYERS = {
    "Ryan Kent": "FFE0B2",
    "Laura Kent": "F8BBD0",
    "Will Cheeseman": "C5E1A5",
    "Daniel Greenhow": "B3E5FC",
    "James Davey": "D1C4E9",
    "Tracey-Anne Greenhow": "FFCCBC",
    "Holly Davey": "F0F4C3",
    "Sophie Maitland": "B2DFDB",
    "Jack Adams": "FFECB3",
    "Sonja Talbot": "E1BEE7",
}

DAYS = [
    ("Friday", "2026-08-14", "Friday 14 Aug"),
    ("Saturday", "2026-08-15", "Saturday 15 Aug"),
    ("Sunday", "2026-08-16", "Sunday 16 Aug"),
]

thin = Side(style="thin", color=RULE)


def load():
    with open(CSV_FILE) as f:
        return list(csv.DictReader(f))


def split_sides(row, player=None):
    """Return (our side, their side). Our side is the one holding a tracked
    player — the given one if named, otherwise whoever the row lists."""
    targets = [player] if player else [p.strip() for p in row["our_players"].split(" / ")]
    for t in targets:
        if t and t in row["side_a"]:
            return row["side_a"], row["side_b"]
        if t and t in row["side_b"]:
            return row["side_b"], row["side_a"]
    return row["side_a"], row["side_b"]


def partner_of(side, player):
    """The other name on a doubles side, or empty for singles."""
    if "&" not in side:
        return ""
    return next((n.strip() for n in side.split("&") if n.strip() != player), "")


def title_block(ws, title, subtitle, width, accent):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BRAND)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name="Calibri", size=10, color=INK)
    c.fill = PatternFill("solid", fgColor=accent)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22


def header_row(ws, headers, row, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(bottom=thin)
    ws.row_dimensions[row].height = 24
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_body(ws, first, last, ncols, court_col, time_col, ours_col, flag_rows=()):
    for r in range(first, last + 1):
        banded = (r - first) % 2 == 1
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            fill = FLAG if r in flag_rows else (BAND if banded else "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.font = Font(name="Calibri", size=10, color=INK)
        ws.cell(row=r, column=time_col).font = Font(name="Calibri", size=11, bold=True, color=INK)
        court = ws.cell(row=r, column=court_col)
        court.font = Font(name="Calibri", size=12, bold=True, color=BRAND)
        court.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=ours_col).font = Font(name="Calibri", size=10, bold=True, color=INK)
        ws.row_dimensions[r].height = 20


def build_overview(wb, rows):
    ws = wb.create_sheet("Overview")
    title_block(ws, "2026 English Open — Squad Schedule", "Proflex English Open · 14–16 August 2026 · all times as published by PickleBook", 6, "F2D5D9")

    header_row(ws, ["Player", "Friday 14 Aug", "Saturday 15 Aug", "Sunday 16 Aug", "Matches", "Events"],
               4, [26, 26, 26, 26, 11, 46])

    by_player = defaultdict(list)
    for r in rows:
        for p in r["our_players"].split(" / "):
            by_player[p.strip()].append(r)

    row = 5
    for player, accent in PLAYERS.items():
        rs = by_player[player]
        ws.cell(row=row, column=1, value=player)
        for i, (day, _, _) in enumerate(DAYS):
            d = [x for x in rs if x["day"] == day]
            if d:
                times = sorted(x["time"] for x in d)
                zones = ", ".join(sorted({x["zone"] for x in d}, key=int))
                val = f"{times[0]}–{times[-1]}   Zone {zones}"
            else:
                val = "—"
            ws.cell(row=row, column=2 + i, value=val)
        ws.cell(row=row, column=5, value=len(rs))
        events = sorted({x["category"].split(" - ", 1)[-1] for x in rs})
        ws.cell(row=row, column=6, value="  ·  ".join(events))

        for c in range(1, 7):
            cell = ws.cell(row=row, column=c)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.font = Font(name="Calibri", size=10, color=INK)
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
        name = ws.cell(row=row, column=1)
        name.font = Font(name="Calibri", size=11, bold=True, color=INK)
        name.fill = PatternFill("solid", fgColor=accent)
        for c in (2, 3, 4):
            if ws.cell(row=row, column=c).value == "—":
                ws.cell(row=row, column=c).font = Font(name="Calibri", size=10, color=MUTED)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=6).font = Font(name="Calibri", size=9, color=MUTED)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    note = ws.cell(row=row, column=1,
                   value="Zones group the courts — check the venue map for where each zone sits. "
                         "Times are scheduled starts and run early or late on the day; the live board at the venue always wins.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    note.font = Font(name="Calibri", size=9, italic=True, color=MUTED)
    note.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.freeze_panes = "A5"
    return ws


def build_day(wb, rows, day, date, label):
    ws = wb.create_sheet(label)
    d = sorted([r for r in rows if r["day"] == day], key=lambda r: (r["time"], int(r["court"])))
    title_block(ws, label, f"{len(d)} matches involving the squad · sorted by start time", 8, "F2D5D9")
    header_row(ws, ["Time", "Court", "Zone", "Event", "Group", "Our players", "Playing", "Match #"],
               4, [9, 9, 8, 30, 10, 34, 34, 10])

    flags = set()
    r = 5
    for row in d:
        ours, theirs = split_sides(row)
        if row["notes"]:
            flags.add(r)
            theirs = f"{theirs}  —  WITHDRAWN"
        ws.cell(row=r, column=1, value=row["time"])
        ws.cell(row=r, column=2, value=int(row["court"]))
        ws.cell(row=r, column=3, value=f"Zone {row['zone']}")
        ws.cell(row=r, column=4, value=row["category"].split(" - ", 1)[-1])
        ws.cell(row=r, column=5, value=row["group"])
        ws.cell(row=r, column=6, value=ours)
        ws.cell(row=r, column=7, value=theirs)
        ws.cell(row=r, column=8, value=f"#{row['match_id']}")
        r += 1

    style_body(ws, 5, r - 1, 8, court_col=2, time_col=1, ours_col=6, flag_rows=flags)
    for rr in range(5, r):
        ws.cell(row=rr, column=3).font = Font(name="Calibri", size=9, color=MUTED)
        ws.cell(row=rr, column=8).font = Font(name="Calibri", size=9, color=MUTED)
    ws.freeze_panes = "A5"
    return ws


def build_player(wb, rows, player, accent):
    ws = wb.create_sheet(player)
    rs = sorted([r for r in rows if player in [p.strip() for p in r["our_players"].split(" / ")]],
                key=lambda r: (r["date"], r["time"]))
    title_block(ws, player, f"{len(rs)} matches · 2026 English Open", 8, accent)
    header_row(ws, ["Day", "Time", "Court", "Zone", "Event", "Group", "Partner", "Playing"],
               4, [11, 9, 9, 8, 30, 10, 24, 34])

    flags = set()
    r = 5
    for row in rs:
        ours, theirs = split_sides(row, player)
        if row["notes"]:
            flags.add(r)
            theirs = f"{theirs}  —  WITHDRAWN"
        ws.cell(row=r, column=1, value=row["day"])
        ws.cell(row=r, column=2, value=row["time"])
        ws.cell(row=r, column=3, value=int(row["court"]))
        ws.cell(row=r, column=4, value=f"Zone {row['zone']}")
        ws.cell(row=r, column=5, value=row["category"].split(" - ", 1)[-1])
        ws.cell(row=r, column=6, value=row["group"])
        ws.cell(row=r, column=7, value=partner_of(ours, player) or "— singles —")
        ws.cell(row=r, column=8, value=theirs)
        r += 1

    style_body(ws, 5, r - 1, 8, court_col=3, time_col=2, ours_col=7, flag_rows=flags)
    for rr in range(5, r):
        ws.cell(row=rr, column=4).font = Font(name="Calibri", size=9, color=MUTED)
        if ws.cell(row=rr, column=7).value == "— singles —":
            ws.cell(row=rr, column=7).font = Font(name="Calibri", size=9, italic=True, color=MUTED)
    ws.freeze_panes = "A5"
    return ws


def build_all_players(wb, rows):
    """Every player in one tab, as colour-banded sections. Used in the compact
    workbook, where ten separate player tabs would not fit the upload budget."""
    ws = wb.create_sheet("By Player")
    title_block(ws, "By Player", "Day · Time · Court · Zone · Event · Partner · Playing — scroll to your name", 7, "F2D5D9")
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 34
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 36

    r = 4
    for player, accent in PLAYERS.items():
        rs = sorted([x for x in rows if player in [p.strip() for p in x["our_players"].split(" / ")]],
                    key=lambda x: (x["date"], x["time"]))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c = ws.cell(row=r, column=1, value=f"{player}   ·   {len(rs)} matches")
        c.font = Font(name="Calibri", size=12, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 26
        r += 1

        first = r
        flags = set()
        for row in rs:
            ours, theirs = split_sides(row, player)
            if row["notes"]:
                flags.add(r)
                theirs = f"{theirs}  —  WITHDRAWN"
            ws.cell(row=r, column=1, value=row["day"])
            ws.cell(row=r, column=2, value=row["time"])
            ws.cell(row=r, column=3, value=int(row["court"]))
            ws.cell(row=r, column=4, value=f"Zone {row['zone']}")
            ws.cell(row=r, column=5, value=f'{row["category"].split(" - ", 1)[-1]}  {row["group"]}')
            ws.cell(row=r, column=6, value=partner_of(ours, player) or "— singles —")
            ws.cell(row=r, column=7, value=theirs)
            r += 1
        style_body(ws, first, r - 1, 7, court_col=3, time_col=2, ours_col=6, flag_rows=flags)
        for rr in range(first, r):
            ws.cell(row=rr, column=4).font = Font(name="Calibri", size=9, color=MUTED)
            if ws.cell(row=rr, column=6).value == "— singles —":
                ws.cell(row=rr, column=6).font = Font(name="Calibri", size=9, italic=True, color=MUTED)
        r += 1
    return ws


def build(rows, per_player_tabs):
    """per_player_tabs builds the full workbook (a tab each for the ten players).
    Without it you get the compact workbook — Overview plus the three day tabs —
    which is small enough to upload to Drive in one piece."""
    wb = Workbook()
    wb.remove(wb.active)
    build_overview(wb, rows)
    for day, date, label in DAYS:
        build_day(wb, rows, day, date, label)
    if per_player_tabs:
        build_all_players(wb, rows)
        for player, accent in PLAYERS.items():
            build_player(wb, rows, player, accent)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    return wb


def main():
    rows = load()
    for name, per_player in ((OUTPUT_FILE, True), (COMPACT_FILE, False)):
        wb = build(rows, per_player)
        wb.save(name)
        print(f"wrote {name}: {len(wb.worksheets)} tabs, {len(rows)} matches")


if __name__ == "__main__":
    main()
